import datetime
import openpyxl

from django.db import transaction
from .models import EmployeeReport, StaffContact


# -------------------------------------------------------
# Convert Persian/Arabic Digits to English Digits
# -------------------------------------------------------
def convert_persian_arabic_digits(value):
    translation_table = str.maketrans(
        "۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩",
        "01234567890123456789",
    )
    return str(value).translate(translation_table)


# -------------------------------------------------------
# Duration Parser
# -------------------------------------------------------
def parse_duration(hhmm: str):
    fallback = {"hours": -1, "minutes": -1}

    if not hhmm or not isinstance(hhmm, str):
        return fallback

    parts = hhmm.strip().split(":")

    if len(parts) != 2:
        return fallback

    try:
        h = int(parts[0])
        m = int(parts[1])
    except ValueError:
        return fallback

    if h < 0 or m < 0 or m > 59:
        return fallback

    return {"hours": h, "minutes": m}


# -------------------------------------------------------
# Normalize Excel Values
# -------------------------------------------------------
def normalize_excel_value(value):
    if value is None:
        return ""

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()

    text = str(value).strip()
    text = convert_persian_arabic_digits(text)

    if text.endswith(".0"):
        text = text[:-2]

    return text


# -------------------------------------------------------
# Normalize National ID
# -------------------------------------------------------
def normalize_national_id(value):
    national_id = normalize_excel_value(value)

    national_id = national_id.replace(" ", "")
    national_id = national_id.replace("-", "")
    national_id = national_id.replace("_", "")

    if national_id.isdigit() and len(national_id) < 10:
        national_id = national_id.zfill(10)

    return national_id


# -------------------------------------------------------
# Normalize Phone Number
# -------------------------------------------------------
def normalize_phone(value):
    phone = normalize_excel_value(value)

    phone = phone.replace(" ", "")
    phone = phone.replace("-", "")
    phone = phone.replace("_", "")
    phone = phone.replace("(", "")
    phone = phone.replace(")", "")
    phone = phone.replace("+", "")

    if phone.startswith("0098"):
        phone = "0" + phone[4:]

    elif phone.startswith("98"):
        phone = "0" + phone[2:]

    elif phone.startswith("9") and len(phone) == 10:
        phone = "0" + phone

    return phone


# -------------------------------------------------------
# Normalize Excel Time
# -------------------------------------------------------
def normalize_time(value):
    if isinstance(value, datetime.time):
        return f"{value.hour:02d}:{value.minute:02d}"

    if value is None:
        return "00:00"

    text = normalize_excel_value(value)

    parts = text.split(":")
    if len(parts) >= 2:
        try:
            return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
        except Exception:
            return "00:00"

    return "00:00"


# -------------------------------------------------------
# Import Employee Reports
# -------------------------------------------------------
def import_excel_reports(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    try:
        sheet = wb.active

        reports_to_create = []
        skipped_rows = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 13:
                skipped_rows += 1
                continue

            national_id = normalize_national_id(row[3])

            if not national_id.isdigit() or len(national_id) != 10:
                skipped_rows += 1
                continue

            try:
                report = EmployeeReport(
                    last_name=normalize_excel_value(row[1]),
                    first_name=normalize_excel_value(row[2]),
                    national_id=national_id,
                    total_presence=normalize_time(row[4]),
                    reduction_work=normalize_time(row[5]),
                    hourly_leave=normalize_time(row[6]),
                    hourly_mission=normalize_time(row[7]),
                    annual_leave_days=int(normalize_excel_value(row[8]) or 0),
                    sick_leave_days=int(normalize_excel_value(row[9]) or 0),
                    daily_mission_days=int(normalize_excel_value(row[10]) or 0),
                    total_overtime=normalize_time(row[11]),
                    total_shift_hours=int(normalize_excel_value(row[12]) or 0),
                )

                reports_to_create.append(report)

            except Exception:
                skipped_rows += 1

        if not reports_to_create:
            return 0

        with transaction.atomic():
            EmployeeReport.objects.all().delete()
            EmployeeReport.objects.bulk_create(reports_to_create, batch_size=1000)

        return EmployeeReport.objects.count()

    finally:
        wb.close()


# -------------------------------------------------------
# Import Staff Contacts
# -------------------------------------------------------
def import_excel_contacts(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    try:
        sheet = wb.active

        contacts_to_create = []
        skipped_rows = []

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 2:
                skipped_rows.append(idx)
                continue

            national_id = normalize_national_id(row[0])
            phone = normalize_phone(row[1])

            if (
                national_id.isdigit()
                and len(national_id) == 10
                and phone.isdigit()
                and len(phone) == 11
                and phone.startswith("09")
            ):
                contacts_to_create.append(
                    StaffContact(
                        national_id=national_id,
                        phone_number=phone,
                    )
                )
            else:
                skipped_rows.append(idx)

        if contacts_to_create:
            created = StaffContact.objects.bulk_create(
                contacts_to_create,
                ignore_conflicts=True,
            )
        else:
            created = []

        return len(created), skipped_rows

    finally:
        wb.close()
